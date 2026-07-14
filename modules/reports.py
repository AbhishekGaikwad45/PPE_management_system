from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file, make_response
from database.db import get_db, fetchall
from datetime import date
import io

reports_bp = Blueprint('reports', __name__)


def get_user_department():
    """Returns None for Admin/Super Admin (sees everything),
    otherwise returns the logged-in user's department so reports get scoped to it."""
    if session.get('role') in ('Admin', 'Super Admin'):
        return None
    return session.get('department')


def get_report_data(c, report_type, from_date, to_date, filter_id=None, department=None):
    if report_type == 'employee':
        c.execute("""SELECT ir.issue_date, e.emp_code, e.name, e.department, e.contractor,
                   i.item_name, i.category, ir.qty, i.unit, ir.status, ir.returnable
                   FROM issue_register ir JOIN employees e ON ir.employee_id=e.id JOIN items i ON ir.item_id=i.id
                   WHERE ir.issue_date BETWEEN %s AND %s
                   AND (%s::int IS NULL OR ir.employee_id = %s::int)
                   AND (%s::text IS NULL OR e.department = %s::text)
                   ORDER BY ir.issue_date DESC""",
                  (from_date, to_date, filter_id, filter_id, department, department))
    elif report_type == 'department':
        c.execute("""SELECT e.department, i.item_name, i.category, SUM(ir.qty) as total_qty, i.unit
                   FROM issue_register ir JOIN employees e ON ir.employee_id=e.id JOIN items i ON ir.item_id=i.id
                   WHERE ir.issue_date BETWEEN %s AND %s
                   AND (%s::text IS NULL OR e.department = %s::text)
                   GROUP BY e.department, i.item_name, i.category, i.unit
                   ORDER BY e.department, total_qty DESC""",
                  (from_date, to_date, department, department))
    elif report_type == 'contractor':
        c.execute("""SELECT e.contractor, i.item_name, i.category, SUM(ir.qty) as total_qty, i.unit
                   FROM issue_register ir JOIN employees e ON ir.employee_id=e.id JOIN items i ON ir.item_id=i.id
                   WHERE ir.issue_date BETWEEN %s AND %s AND e.contractor IS NOT NULL AND e.contractor != ''
                   GROUP BY e.contractor, i.item_name, i.category, i.unit ORDER BY e.contractor, total_qty DESC""", (from_date, to_date))
    elif report_type == 'item':
        c.execute("""SELECT i.item_code, i.item_name, i.category, SUM(ir.qty) as total_issued, i.unit, i.stock
                   FROM issue_register ir JOIN items i ON ir.item_id=i.id
                   WHERE ir.issue_date BETWEEN %s AND %s GROUP BY i.item_code, i.item_name, i.category, i.unit, i.stock
                   ORDER BY total_issued DESC""", (from_date, to_date))
    elif report_type == 'monthly':
        c.execute("""SELECT TO_CHAR(TO_DATE(ir.issue_date,'YYYY-MM-DD'),'YYYY-MM') as month,
                   i.item_name, i.category, SUM(ir.qty) as total_qty
                   FROM issue_register ir JOIN items i ON ir.item_id=i.id
                   WHERE ir.issue_date BETWEEN %s AND %s GROUP BY month, i.item_name, i.category ORDER BY month DESC""", (from_date, to_date))
    elif report_type == 'stock':
        c.execute("""SELECT item_code, item_name, category, unit, stock, min_stock,
                   CASE WHEN stock <= min_stock THEN 'LOW STOCK' ELSE 'OK' END as status
                   FROM items ORDER BY category, item_name""")
    return fetchall(c)


@reports_bp.route('/reports')
def index():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    department = get_user_department()
    conn = get_db(); c = conn.cursor()
    if department:
        c.execute("SELECT id, emp_code, name FROM employees WHERE department=%s ORDER BY name", (department,))
    else:
        c.execute("SELECT id, emp_code, name FROM employees ORDER BY name")
    employees = fetchall(c)
    conn.close()
    today = date.today()
    from_date = date(today.year, today.month, 1).strftime('%Y-%m-%d')
    to_date = today.strftime('%Y-%m-%d')
    return render_template('reports.html', employees=employees, from_date=from_date, to_date=to_date,
                           department=department)


@reports_bp.route('/reports/view', methods=['GET', 'POST'])
def view():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    department = get_user_department()
    report_type = request.form.get('report_type') or request.args.get('report_type', 'employee')
    from_date = request.form.get('from_date') or request.args.get('from_date', date.today().strftime('%Y-01-01'))
    to_date = request.form.get('to_date') or request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    filter_id = request.form.get('filter_id')

    if filter_id is None:
        filter_id = request.args.get('filter_id')

    if filter_id == "":
        filter_id = None
    elif filter_id is not None:
        filter_id = int(filter_id)
    conn = get_db(); c = conn.cursor()
    data = get_report_data(c, report_type, from_date, to_date, filter_id, department)
    if department:
        c.execute("SELECT id, emp_code, name FROM employees WHERE department=%s ORDER BY name", (department,))
    else:
        c.execute("SELECT id, emp_code, name FROM employees ORDER BY name")
    employees = fetchall(c)
    conn.close()
    return render_template('report_view.html', data=data, report_type=report_type,
                           from_date=from_date, to_date=to_date, employees=employees, filter_id=filter_id)


@reports_bp.route('/reports/export/excel', methods=['GET', 'POST'])
def export_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Install openpyxl: pip install openpyxl', 'danger')
        return redirect(url_for('reports.index'))
    department = get_user_department()
    report_type = request.args.get('report_type', 'employee')
    from_date = request.args.get('from_date', date.today().strftime('%Y-01-01'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    filter_id = request.args.get('filter_id')

    if filter_id == "":
        filter_id = None
    elif filter_id is not None:
        filter_id = int(filter_id)
    conn = get_db(); c = conn.cursor()
    data = get_report_data(c, report_type, from_date, to_date, filter_id, department)
    conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = report_type.upper()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill('solid', fgColor='1A3A5C')
    center = Alignment(horizontal='center')
    ws.merge_cells('A1:H1')
    ws['A1'].value = f'PPE & Equipment Management System - {report_type.upper()} REPORT'
    ws['A1'].font = Font(bold=True, size=14, color='1A3A5C'); ws['A1'].alignment = center
    ws.merge_cells('A2:H2')
    ws['A2'].value = f'Period: {from_date} to {to_date}'; ws['A2'].alignment = center
    if data:
        headers = list(data[0].keys())
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h.upper().replace('_', ' '))
            cell.font = header_font; cell.fill = header_fill; cell.alignment = center
        alt_fill = PatternFill('solid', fgColor='E8F0F7')
        for row_idx, row in enumerate(data, 5):
            for col_idx, key in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key))
                if row_idx % 2 == 0: cell.fill = alt_fill
        from openpyxl.utils import get_column_letter

        for column_cells in ws.iter_cols(
                min_col=1,
                max_col=ws.max_column,
                min_row=4,
                max_row=ws.max_row):

            max_length = 0

            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[column].width = min(max_length + 4, 40)
    output = io.BytesIO(); wb.save(output); output.seek(0)
    return send_file(output, as_attachment=True, download_name=f'PPE_{report_type}_{from_date}_{to_date}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/reports/export/pdf')
def export_pdf():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    department = get_user_department()
    report_type = request.args.get('report_type', 'employee')
    from_date = request.args.get('from_date', date.today().strftime('%Y-01-01'))
    to_date = request.args.get('to_date', date.today().strftime('%Y-%m-%d'))
    filter_id = request.args.get('filter_id')

    if filter_id == "":
        filter_id = None
    elif filter_id is not None:
        filter_id = int(filter_id)
    conn = get_db(); c = conn.cursor()
    data = get_report_data(c, report_type, from_date, to_date, filter_id, department)
    conn.close()
    headers = list(data[0].keys()) if data else []
    html = f"""<!DOCTYPE html><html><head><style>
    body{{font-family:Arial;font-size:10px;margin:20px}}
    h2{{color:#1A3A5C;text-align:center}} p{{text-align:center;color:#666}}
    table{{width:100%;border-collapse:collapse;margin-top:10px}}
    th{{background:#1A3A5C;color:white;padding:6px;text-align:left}}
    td{{padding:5px;border-bottom:1px solid #ddd}} tr:nth-child(even){{background:#f0f6ff}}
    </style></head><body>
    <h2>PPE & Equipment Management System</h2>
    <p>{report_type.upper()} REPORT | {from_date} to {to_date}</p>
    <table><tr>{''.join(f'<th>{h.upper().replace("_", " ")}</th>' for h in headers)}</tr>
    {''.join('<tr>' + ''.join(f'<td>{row.get(k, "")}</td>' for k in headers) + '</tr>' for row in data)}
    </table></body></html>"""
    response = make_response(html)
    response.headers['Content-Type'] = 'text/html'
    return response