from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, Response
from database.db import get_db, fetchall, fetchone
from modules.user_admin import has_permission   # ← ADD — reuse the same role_permissions table
import io

employees_bp = Blueprint('employees', __name__)

PER_PAGE = 100

# Some department names in your real data are variants of the same thing.
# Group them under one combined display card. Match is case-insensitive.
COMBINE_GROUPS = {
    'Civil & Project': [
        'Civil', 'Civil/Project', 'Civil & Project', 'Civil and Project',
        'Project', 'Projects'                      # ← ADDED 'Projects' (plural)
    ],
    'Administration / HR Admin': [
        'Administration', 'HR/Admin', 'HR & Admin', 'HR and Admin', 'HR',
        'Admin'                                     # ← ADDED 'Admin' (short form)
    ],
    'IT / Information Technology': [
        'IT', 'Information Technology',
        'I.T.', 'Information Tech'                  # ← ADDED common IT variants too, just in case
    ],
}
# reverse lookup: UPPER(raw name) -> display name
_COMBINE_LOOKUP = {variant.upper(): display for display, variants in COMBINE_GROUPS.items() for variant in variants}


def _display_dept_name(raw_dept):
    """Map a raw department string to its combined display name, if it's part of a group."""
    if not raw_dept:
        return 'Unassigned'
    return _COMBINE_LOOKUP.get(raw_dept.strip().upper(), raw_dept.strip())


@employees_bp.route('/employees')
def index():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db()
    c = conn.cursor()
    role = session.get('role')
    dept = session.get('department')
    is_admin = role in ['Admin', 'Super Admin']

    # Optional filters from query string (?department=X or ?contractor=X, &page=N)
    dept_filter = request.args.get('department') or ''
    contractor_filter = request.args.get('contractor') or ''
    page = request.args.get('page', 1, type=int)
    if page < 1:
        page = 1
    offset = (page - 1) * PER_PAGE

    # Pull raw department text for every Active employee and group in Python —
    # more reliable than SQL TRIM(), which does not strip \r / \t / other
    # hidden whitespace that can sneak in from the SQL Server import.
    c.execute("SELECT department FROM employees WHERE status='Active'")
    raw_dept_rows = fetchall(c)

    dept_counts = {}
    display_to_raw = {}
    for row in raw_dept_rows:
        raw = row['department']
        raw_clean = raw.strip() if raw else ''
        display = _display_dept_name(raw_clean) if raw_clean else 'Unassigned'
        dept_counts[display] = dept_counts.get(display, 0) + 1
        display_to_raw.setdefault(display, set()).add(raw)  # keep the exact original raw value for filtering
    display_to_raw = {k: list(v) for k, v in display_to_raw.items()}

    department_names = sorted(d for d in dept_counts if d != 'Unassigned')

    # Build WHERE clause once, reuse for both COUNT and the paginated SELECT
    where_clauses = ["status='Active'"]
    params = []
    if not is_admin:
            display_name = _display_dept_name(dept)
            raw_values = display_to_raw.get(display_name, [dept])
            placeholders = ','.join(['%s'] * len(raw_values))
            where_clauses.append(f"department IN ({placeholders})")
            params.extend(raw_values)
    elif contractor_filter:
        where_clauses.append("contractor=%s")
        params.append(contractor_filter)
    elif dept_filter:
        if dept_filter == 'Unassigned':
            where_clauses.append("(department IS NULL OR TRIM(department)='')")
        else:
            raw_values = [r for r in display_to_raw.get(dept_filter, [dept_filter]) if r]
            if raw_values:
                placeholders = ','.join(['%s'] * len(raw_values))
                where_clauses.append(f"department IN ({placeholders})")
                params.extend(raw_values)
    where_sql = "WHERE " + " AND ".join(where_clauses)

    # Total count for this filter (fast — COUNT only, no row data)
    c.execute(f"SELECT COUNT(*) AS n FROM employees {where_sql}", params)
    total = fetchone(c)['n']
    total_pages = max(1, (total + PER_PAGE - 1) // PER_PAGE)
    if page > total_pages:
        page = total_pages
        offset = (page - 1) * PER_PAGE

    # Only fetch ONE page of rows, not all 5000+
    c.execute(
        f"SELECT * FROM employees {where_sql} ORDER BY department NULLS LAST, name LIMIT %s OFFSET %s",
        params + [PER_PAGE, offset]
    )
    employees = fetchall(c)

    if not is_admin:
        dept_counts = {}
        department_names = []

    # Departments (used for Add/Edit dropdowns — unaffected by the combine-groups above)
    if is_admin:
        c.execute("SELECT name FROM departments ORDER BY name")
    else:
        c.execute("SELECT name FROM departments WHERE name=%s", (dept,))
    departments = fetchall(c)

    # Contractors
    c.execute("SELECT name FROM contractors ORDER BY name")
    contractors = fetchall(c)
    conn.close()

    can_create = has_permission('can_create')
    can_edit = has_permission('can_edit')
    can_delete = has_permission('can_delete')

    return render_template('employees.html', employees=employees, departments=departments,
                            department_names=department_names,
                            contractors=contractors, can_create=can_create,
                            can_edit=can_edit, can_delete=can_delete,
                            is_admin=is_admin, dept_counts=dept_counts,
                            dept_filter=dept_filter, contractor_filter=contractor_filter,
                            page=page, total_pages=total_pages,
                            total=total, per_page=PER_PAGE)

@employees_bp.route('/employees/add', methods=['POST'])
def add():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_create'):                       # ← ADD
        flash('You do not have permission to add employees.', 'danger')
        return redirect(url_for('employees.index'))
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("INSERT INTO employees (emp_code,name,department,contractor,designation,status) VALUES (%s,%s,%s,%s,%s,%s)",
                  (request.form['emp_code'], request.form['name'], request.form['department'],
                   request.form.get('contractor',''), request.form.get('designation',''), 'Active'))
        conn.commit()
        flash('Employee added successfully.', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('employees.index'))

@employees_bp.route('/employees/edit/<int:id>', methods=['POST'])
def edit(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if not has_permission('can_edit'):
        flash('You do not have permission to edit employees.', 'danger')
        return redirect(url_for('employees.index'))

    conn = get_db()
    c = conn.cursor()

    try:
        c.execute("SELECT department FROM employees WHERE id=%s", (id,))
        old = fetchone(c)

        department = request.form.get("department")
        if not department:
            department = old["department"]
            print("Contractor =", request.form.get("contractor"))
            print(request.form)
        c.execute("""
            UPDATE employees
            SET
                emp_code=%s,
                name=%s,
                department=%s,
                contractor=%s,
                designation=%s,
                status=%s
            WHERE id=%s
        """,
        (
            request.form.get("emp_code"),
            request.form.get("name"),
            department,
            request.form.get("contractor"),
            request.form.get("designation"),
            request.form.get("status"),
            id
        ))

        conn.commit()
        flash("Employee updated successfully.", "success")

    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")

    finally:
        conn.close()

    return redirect(url_for('employees.index'))

@employees_bp.route('/employees/delete/<int:id>')
def delete(id):
    if 'user' not in session or not has_permission('can_delete'):   # ← CHANGED — was Admin/Super Admin only
        flash('Access denied.', 'danger')
        return redirect(url_for('employees.index'))
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM employees WHERE id=%s", (id,))
    conn.commit()
    conn.close()
    flash('Employee deleted.', 'info')
    return redirect(url_for('employees.index'))

@employees_bp.route('/employees/search')
def search():
    q = request.args.get('q','')
    conn = get_db()
    c = conn.cursor()
    dept = session.get('department')
    if dept:
        c.execute("SELECT id, emp_code, name, department, contractor FROM employees WHERE (name ILIKE %s OR emp_code ILIKE %s) AND department=%s AND status='Active' LIMIT 20",
                  (f'%{q}%', f'%{q}%', dept))
    else:
        c.execute("SELECT id, emp_code, name, department, contractor FROM employees WHERE (name ILIKE %s OR emp_code ILIKE %s) AND status='Active' LIMIT 20",
                  (f'%{q}%', f'%{q}%'))
    results = fetchall(c)
    conn.close()
    return jsonify(results)

@employees_bp.route('/employees/export-excel')
def export_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl is not installed.', 'danger')
        return redirect(url_for('employees.index'))
    conn = get_db()
    c = conn.cursor()
    dept = session.get('department')
    if dept:
        c.execute("SELECT * FROM employees WHERE department=%s ORDER BY name", (dept,))
    else:
        c.execute("SELECT * FROM employees ORDER BY name")
    employees = fetchall(c)
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1B3A5C")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    headers = ["Emp Code", "Name", "Department", "Contractor", "Designation", "Status"]
    col_widths = [14, 28, 22, 22, 22, 12]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width
    active_fill = PatternFill("solid", fgColor="E8F5E9")
    inactive_fill = PatternFill("solid", fgColor="FAFAFA")
    for row_idx, emp in enumerate(employees, start=2):
        row_data = [emp.get("emp_code"), emp.get("name"), emp.get("department",""),
                    emp.get("contractor",""), emp.get("designation",""), emp.get("status")]
        fill = active_fill if emp.get("status") == "Active" else inactive_fill
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill; cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return Response(output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employees.xlsx"})

@employees_bp.route('/employees/import-excel', methods=['POST'])
def import_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_create'):
        flash('Access denied.', 'danger')
        return redirect(url_for('employees.index'))
    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('employees.index'))
    file = request.files['excel_file']
    try:
        import openpyxl
    except ImportError:
        flash('openpyxl not installed.', 'danger')
        return redirect(url_for('employees.index'))
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        HEADER_MAP = {'emp code':'emp_code','empcode':'emp_code','employee code':'emp_code','emp_code':'emp_code','code':'emp_code',
                      'name':'name','employee name':'name','full name':'name','department':'department','dept':'department',
                      'contractor':'contractor','designation':'designation','position':'designation','status':'status'}
        first_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {}
        for idx, cell_val in enumerate(first_row):
            if cell_val is None: continue
            key = str(cell_val).strip().lower()
            if key in HEADER_MAP and HEADER_MAP[key] not in col_map:
                col_map[HEADER_MAP[key]] = idx
        has_headers = 'emp_code' in col_map and 'name' in col_map
        start_row = 2 if has_headers else 1
        POSITIONAL = {'emp_code':0,'name':1,'department':2,'contractor':3,'designation':4,'status':5}

        def get_val(row_values, field):
            mapping = col_map if has_headers else POSITIONAL
            idx = mapping.get(field)
            if idx is None or idx >= len(row_values): return ''
            v = row_values[idx]
            return str(v).strip() if v is not None else ''

        conn = get_db(); c = conn.cursor()
        added = updated = skipped = unchanged = 0
        UPDATABLE_FIELDS = ['name', 'department', 'contractor', 'designation', 'status']

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not any(row): continue
            row_values = list(row)
            emp_code = get_val(row_values, 'emp_code')
            name = get_val(row_values, 'name')

            # emp_code is still required to identify the row — name is NOT checked for duplicates
            if not emp_code:
                skipped += 1
                continue

            status = get_val(row_values, 'status') or 'Active'
            if status not in ('Active', 'Inactive'):
                status = 'Active'

            incoming = {
                'name': name,
                'department': get_val(row_values, 'department'),
                'contractor': get_val(row_values, 'contractor'),
                'designation': get_val(row_values, 'designation'),
                'status': status,
            }

            try:
                # Check if this emp_code already exists
                c.execute("SELECT * FROM employees WHERE emp_code=%s", (emp_code,))
                existing = fetchone(c)

                if existing is None:
                    # New employee — insert (name required only for new rows)
                    if not name:
                        skipped += 1
                        continue
                    c.execute("""
                        INSERT INTO employees (emp_code, name, department, contractor, designation, status)
                        VALUES (%s, %s, %s, %s, %s, %s)
                    """, (emp_code, incoming['name'], incoming['department'],
                          incoming['contractor'], incoming['designation'], incoming['status']))
                    added += 1
                else:
                    # Existing employee — build UPDATE only for fields that actually changed
                    # blank incoming values don't overwrite existing data
                    changed_fields = {}
                    for field in UPDATABLE_FIELDS:
                        new_val = incoming[field]
                        old_val = existing.get(field) or ''
                        if new_val and new_val != old_val:
                            changed_fields[field] = new_val

                    if changed_fields:
                        set_clause = ", ".join(f"{f}=%s" for f in changed_fields)
                        params = list(changed_fields.values()) + [emp_code]
                        c.execute(f"UPDATE employees SET {set_clause} WHERE emp_code=%s", params)
                        updated += 1
                    else:
                        unchanged += 1

            except Exception:
                skipped += 1

        conn.commit(); conn.close()
        flash(f'Import complete: {added} added, {updated} updated, {unchanged} unchanged, {skipped} skipped.', 'success')
    except Exception as e:
        flash(f'Failed to read Excel: {e}', 'danger')
    return redirect(url_for('employees.index'))

@employees_bp.route('/employees/by-contractor')
def by_contractor_partial():
    """Returns an HTML fragment (table + edit modals) for one contractor's
    Active employees. Loaded via AJAX only when that contractor's popup is
    opened — keeps the Contractors page itself light."""
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    name = request.args.get('name', '')
    if not name:
        return '<div class="p-3 text-muted">No contractor specified.</div>'

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM employees WHERE contractor=%s AND status='Active' ORDER BY name", (name,))
    employees = fetchall(c)
    c.execute("SELECT name FROM departments ORDER BY name")
    departments = fetchall(c)
    c.execute("SELECT name FROM contractors ORDER BY name")
    contractors = fetchall(c)
    conn.close()

    can_edit = has_permission('can_edit')
    can_delete = has_permission('can_delete')

    return render_template('contractor_employees_table.html',
                            employees=employees, departments=departments, contractors=contractors,
                            can_edit=can_edit, can_delete=can_delete)


@employees_bp.route('/contractors')
def contractors():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    role = session.get('role')
    dept = session.get('department')
    is_admin = role in ['Admin', 'Super Admin']

    if is_admin:
        c.execute("SELECT * FROM contractors ORDER BY department NULLS LAST, name")
        contractors = fetchall(c)
        c.execute("SELECT name FROM departments ORDER BY name")
        departments = fetchall(c)
    else:
        c.execute("SELECT * FROM contractors WHERE department=%s ORDER BY name", (dept,))
        contractors = fetchall(c)
        departments = [{'name': dept}] if dept else []

    conn.close()

    can_create = has_permission('can_create')
    can_edit = has_permission('can_edit')
    can_delete = has_permission('can_delete')

    # Group contractors by department for the admin popup view
    contractors_by_dept = {}
    for ctr in contractors:
        d = ctr.get('department') or 'Unassigned'
        contractors_by_dept.setdefault(d, []).append(ctr)

    # Employee count per contractor (Active employees only) — for the new
    # "Employees by Contractor" card grid
    c2 = get_db()
    cc = c2.cursor()
    cc.execute("SELECT contractor FROM employees WHERE status='Active' AND contractor IS NOT NULL AND TRIM(contractor) != ''")
    contractor_emp_counts = {}
    for row in fetchall(cc):
        name = row['contractor'].strip()
        contractor_emp_counts[name] = contractor_emp_counts.get(name, 0) + 1
    c2.close()

    return render_template('contractors.html', contractors=contractors, departments=departments,
                            contractors_by_dept=contractors_by_dept, is_admin=is_admin,
                            contractor_emp_counts=contractor_emp_counts,
                            can_create=can_create, can_edit=can_edit, can_delete=can_delete)

@employees_bp.route('/contractors/add', methods=['POST'])
def add_contractor():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_create'):
        flash('You do not have permission to add contractors.', 'danger')
        return redirect(url_for('employees.contractors'))

    role = session.get('role')
    dept = session.get('department')
    is_admin = role in ['Admin', 'Super Admin']

    # Admin picks the department from the form; non-admin is locked to their own department
    department = request.form.get('department') if is_admin else dept

    conn = get_db(); c = conn.cursor()
    try:
        c.execute("INSERT INTO contractors (name, contact, department) VALUES (%s,%s,%s)",
                  (request.form['name'], request.form.get('contact',''), department))
        conn.commit()
        flash('Contractor added.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('employees.contractors'))

@employees_bp.route('/contractors/edit/<int:contractor_id>', methods=['POST'])
def edit_contractor(contractor_id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_edit'):
        flash('You do not have permission to edit contractors.', 'danger')
        return redirect(url_for('employees.contractors'))

    role = session.get('role')
    dept = session.get('department')
    is_admin = role in ['Admin', 'Super Admin']

    conn = get_db(); c = conn.cursor()
    try:
        c.execute("SELECT department FROM contractors WHERE id=%s", (contractor_id,))
        existing = fetchone(c)

        if is_admin:
            department = request.form.get('department') or (existing['department'] if existing else None)
        else:
            # Non-admin can only edit contractors already in their own department, and can't move them elsewhere
            department = dept

        c.execute("UPDATE contractors SET name=%s, contact=%s, department=%s WHERE id=%s",
                  (request.form['name'], request.form.get('contact',''), department, contractor_id))
        conn.commit()
        flash('Contractor updated.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('employees.contractors'))


@employees_bp.route('/contractors/delete/<int:contractor_id>', methods=['POST'])
def delete_contractor(contractor_id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_delete'):                        # ← ADD
        flash('You do not have permission to delete contractors.', 'danger')
        return redirect(url_for('employees.contractors'))
    conn = get_db(); c = conn.cursor()
    try:
        c.execute("DELETE FROM contractors WHERE id=%s", (contractor_id,))
        conn.commit()
        flash('Contractor deleted.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('employees.contractors'))

# ───────────────────────── EXCEL EXPORT — Contractors ─────────────────────────
@employees_bp.route('/contractors/export-excel')
def export_contractors_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl is not installed.', 'danger')
        return redirect(url_for('employees.contractors'))

    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM contractors ORDER BY name")
    contractors_list = fetchall(c)
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contractors"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Contractor Name", "Contact", "Department"]
    col_widths = [30, 22, 22]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, ctr in enumerate(contractors_list, start=2):
        row_data = [ctr.get("name"), ctr.get("contact", ""), ctr.get("department", "")]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return Response(output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contractors.xlsx"})


# ───────────────────────── EXCEL IMPORT — Contractors (upsert + validation) ─────────────────────────
@employees_bp.route('/contractors/import-excel', methods=['POST'])
def import_contractors_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_create'):
        flash('You do not have permission to import contractors.', 'danger')
        return redirect(url_for('employees.contractors'))

    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('employees.contractors'))

    file = request.files['excel_file']
    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash('Invalid file type. Please upload a .xlsx or .xls file.', 'danger')
        return redirect(url_for('employees.contractors'))

    try:
        import openpyxl
    except ImportError:
        flash('openpyxl not installed.', 'danger')
        return redirect(url_for('employees.contractors'))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Header names matched case-insensitively / regardless of spacing
        HEADER_MAP = {
            'contractor name': 'name', 'contractorname': 'name', 'name': 'name',
            'contact': 'contact', 'contact number': 'contact', 'phone': 'contact',
            'contact no': 'contact', 'mobile': 'contact',
            'department': 'department', 'dept': 'department',
        }

        first_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {}
        for idx, cell_val in enumerate(first_row):
            if cell_val is None:
                continue
            key = str(cell_val).strip().lower()
            if key in HEADER_MAP and HEADER_MAP[key] not in col_map:
                col_map[HEADER_MAP[key]] = idx

        has_headers = 'name' in col_map
        start_row = 2 if has_headers else 1
        POSITIONAL = {'name': 0, 'contact': 1, 'department': 2}

        def get_val(row_values, field):
            mapping = col_map if has_headers else POSITIONAL
            idx = mapping.get(field)
            if idx is None or idx >= len(row_values):
                return ''
            v = row_values[idx]
            return str(v).strip() if v is not None else ''

        conn = get_db(); c = conn.cursor()
        added = updated = unchanged = skipped = 0
        error_log = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not any(row):
                continue
            row_values = list(row)

            name = get_val(row_values, 'name')
            contact = get_val(row_values, 'contact')
            department = get_val(row_values, 'department')

            # Contractor Name is the required key — matched case-insensitively against existing rows
            if not name:
                skipped += 1
                error_log.append(f"Row {row_idx}: missing Contractor Name — skipped")
                continue

            try:
                c.execute("SELECT * FROM contractors WHERE LOWER(name)=LOWER(%s)", (name,))
                existing = fetchone(c)

                if existing is None:
                    c.execute("INSERT INTO contractors (name, contact, department) VALUES (%s, %s, %s)",
                              (name, contact, department))
                    added += 1
                else:
                    changed_fields = {}
                    if name != existing.get('name'):
                        changed_fields['name'] = name
                    old_contact = existing.get('contact') or ''
                    if contact and contact != old_contact:
                        changed_fields['contact'] = contact
                    old_department = existing.get('department') or ''
                    if department and department != old_department:
                        changed_fields['department'] = department

                    if changed_fields:
                        set_clause = ", ".join(f"{f}=%s" for f in changed_fields)
                        params = list(changed_fields.values()) + [existing['id']]
                        c.execute(f"UPDATE contractors SET {set_clause} WHERE id=%s", params)
                        updated += 1
                    else:
                        unchanged += 1

            except Exception as row_err:
                skipped += 1
                error_log.append(f"Row {row_idx}: {row_err}")

        conn.commit(); conn.close()

        summary = f'Import complete: {added} added, {updated} updated, {unchanged} unchanged, {skipped} skipped.'
        flash(summary, 'success')
        if error_log:
            flash('Issues found: ' + ' | '.join(error_log[:10]) + (' ...' if len(error_log) > 10 else ''), 'warning')

    except Exception as e:
        flash(f'Failed to read Excel: {e}', 'danger')

    return redirect(url_for('employees.contractors'))