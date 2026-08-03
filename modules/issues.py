from flask import Blueprint, render_template, request, redirect, url_for, session, flash, Response
from database.db import get_db, fetchall, fetchone
from datetime import date
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from modules.employees import _display_dept_name, COMBINE_GROUPS
from modules.user_admin import has_permission   # ← ADDED

issues_bp = Blueprint('issues', __name__)
_thin = Side(style='thin', color='C7CDD4')
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left', vertical='center', wrap_text=True)


@issues_bp.route('/issues')
def index():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    # View access — everyone with login can see their own dept issues; no gate needed here
    conn = get_db(); c = conn.cursor()
    dept = session.get('department')
    role = session.get('role')
    is_admin = role in ['Admin', 'Super Admin']

    if not is_admin and dept:
        display_name = _display_dept_name(dept)
        dept_variants = [v.lower() for v in COMBINE_GROUPS.get(display_name, [dept])]
    else:
        dept_variants = []

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    if dept_variants:
        c.execute("SELECT id, emp_code, name, department FROM employees WHERE status='Active' AND LOWER(department)=ANY(%s) ORDER BY name", (dept_variants,))
    else:
        c.execute("SELECT id, emp_code, name, department FROM employees WHERE status='Active' ORDER BY name")
    employees_raw = fetchall(c)
    employees = [{"id": e["id"], "label": f"{e['emp_code']} - {e['name']} ({e['department']})", "badge": f"{e['emp_code']} — {e['name']}"} for e in employees_raw]

    if is_admin or not dept_variants:
        c.execute("SELECT id, item_name, stock, unit FROM items ORDER BY item_name")
        items_raw = fetchall(c)
        items = [{"id": i["id"], "stock": i["stock"], "unit": i["unit"], "label": f"{i['item_name']} [Stock: {i['stock']} {i['unit']}]", "badge": f"{i['item_name']} [{i['stock']} {i['unit']}]"} for i in items_raw]
    else:
        # Calculate real department-specific stock balance for each item
        c.execute("""
            SELECT 
                i.id, i.item_name, i.unit, i.added_by_department,
                (
                    COALESCE((
                        SELECT SUM(r.qty) 
                        FROM stock_receipts r 
                        WHERE r.item_id = i.id AND LOWER(r.department) = ANY(%s)
                    ), 0)
                    -
                    COALESCE((
                        SELECT SUM(ir.qty) 
                        FROM issue_register ir 
                        WHERE ir.item_id = i.id AND LOWER(ir.department) = ANY(%s)
                    ), 0)
                    +
                    COALESCE((
                        SELECT SUM(rr.qty_no) 
                        FROM return_register rr 
                        JOIN issue_register ir2 ON ir2.id = rr.issue_id 
                        WHERE rr.item_id = i.id AND LOWER(ir2.department) = ANY(%s)
                    ), 0)
                ) AS dept_stock
            FROM items i
            ORDER BY i.item_name
        """, (dept_variants, dept_variants, dept_variants))
        items_raw = fetchall(c)

        items = []
        dept_lower = (dept or '').strip().lower()
        for i in items_raw:
            d_stock = i["dept_stock"]
            added_by = (i["added_by_department"] or '').strip().lower()

            # Show item if added by this department, or has department stock > 0, or is global (added_by == '')
            if added_by == dept_lower or d_stock > 0 or added_by == '':
                items.append({
                    "id": i["id"],
                    "stock": d_stock,
                    "unit": i["unit"],
                    "label": f"{i['item_name']} [Stock: {d_stock} {i['unit']}]",
                    "badge": f"{i['item_name']} [{d_stock} {i['unit']}]"
                })

    query = """
        SELECT ir.*, e.name as emp_name, e.emp_code, e.department, i.item_name, i.unit
        FROM issue_register ir
        JOIN employees e ON ir.employee_id=e.id
        JOIN items i ON ir.item_id=i.id
        WHERE 1=1
    """
    params = []
    if dept_variants:
        query += " AND LOWER(e.department)=ANY(%s)"
        params.append(dept_variants)
    if from_date:
        query += " AND ir.issue_date >= %s"
        params.append(from_date)
    if to_date:
        query += " AND ir.issue_date <= %s"
        params.append(to_date)
    query += " ORDER BY ir.issue_date DESC"

    c.execute(query, tuple(params))
    issues = fetchall(c)
    conn.close()

    # ← FIXED — has_permission() takes ONE argument ('can_create' / 'can_edit'
    # / 'can_delete'), not (module, action). The old two-arg call silently
    # always evaluated wrong, which is why Add/Edit/Delete stayed visible
    # even for roles with no permission.
    can_create = has_permission('can_create')
    can_edit   = has_permission('can_edit')
    can_delete = has_permission('can_delete')

    return render_template('issues.html', employees=employees, items=items, issues=issues,
                            today=date.today(), from_date=from_date, to_date=to_date,
                            can_create=can_create, can_edit=can_edit, can_delete=can_delete)


def find_issue_department(cur, item_id, employee_department, qty):
    """
    Return the department from which stock should be issued.
    Loops through the combined-department group for the employee's
    department and returns the first one whose ledger balance
    (receipts - issues + returns) can cover `qty`. Returns None if
    no department in the group has enough stock.
    """

    from modules.employees import COMBINE_GROUPS, _display_dept_name

    display_name = _display_dept_name(employee_department)

    departments = COMBINE_GROUPS.get(display_name, [employee_department])

    for dept in departments:

        cur.execute("""
            SELECT
                COALESCE((
                    SELECT SUM(qty)
                    FROM stock_receipts
                    WHERE item_id=%s
                      AND LOWER(department)=LOWER(%s)
                ),0)

                -

                COALESCE((
                    SELECT SUM(qty)
                    FROM issue_register
                    WHERE item_id=%s
                      AND LOWER(department)=LOWER(%s)
                ),0)

                +

                COALESCE((
                    SELECT SUM(r.qty_no)
                    FROM return_register r
                    JOIN issue_register i
                      ON i.id = r.issue_id
                    WHERE r.item_id=%s
                      AND LOWER(i.department)=LOWER(%s)
                ),0)

                AS balance
        """, (
            item_id, dept,
            item_id, dept,
            item_id, dept
        ))

        row = fetchone(cur)

        if row and row["balance"] >= qty:
            return dept

    return None


@issues_bp.route('/issues/add', methods=['POST'])
def add():
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    role = session.get("role")
    if not has_permission('can_create'):                       # ← FIXED
        flash("You don't have permission to issue PPE/Equipment.", "danger")
        return redirect(url_for("issues.index"))

    conn = get_db()
    c = conn.cursor()
    issued_count = 0

    try:
        employee_ids = [int(x) for x in request.form['employee_id'].split(',') if x.strip()]
        item_ids = [int(x) for x in request.form['item_id'].split(',') if x.strip()]

        qty = int(request.form['qty'])
        issue_date = request.form['issue_date']
        remarks = request.form.get('remarks', '')
        issued_by = session['full_name']

        returnable = 1 if request.form.get('returnable') else 0
        return_due = request.form.get('return_due_date') if returnable else None

        dept = session.get("department")
        is_admin = role in ["Admin", "Super Admin"]

        if not is_admin:
            display_name = _display_dept_name(dept)
            allowed_variants = [v.lower() for v in COMBINE_GROUPS.get(display_name, [dept])]
            for emp_id in employee_ids:
                c.execute("SELECT department FROM employees WHERE id=%s", (emp_id,))
                emp = fetchone(c)
                if not emp or (emp["department"] or '').lower() not in allowed_variants:
                    conn.close()
                    flash("You can only issue PPE to employees in your department.", "danger")
                    return redirect(url_for("issues.index"))

        for item_id in item_ids:
            c.execute("SELECT stock FROM items WHERE id=%s", (item_id,))
            stock = fetchone(c)
            if not stock or stock["stock"] < qty:
                conn.close()
                flash("Insufficient stock!", "danger")
                return redirect(url_for("issues.index"))

        for emp_id in employee_ids:

            c.execute("SELECT department FROM employees WHERE id=%s", (emp_id,))
            emp_row = fetchone(c)

            employee_department = emp_row["department"] if emp_row else dept

            for item_id in item_ids:

                # ← FIXED — get_department_available_stock() was never
                # defined anywhere in this file, so every call here raised
                # a NameError, was swallowed by the except block below, and
                # silently rolled back the whole issue. Replaced with the
                # existing find_issue_department() helper, which already
                # loops through the employee's COMBINE_GROUPS departments
                # and returns the first one with enough ledger balance.
                issue_department = find_issue_department(
                    c, item_id, employee_department, qty
                )

                if not issue_department:
                    conn.rollback()
                    flash(
                        f"Insufficient stock in '{employee_department}' "
                        f"(or related departments) for this item.",
                        "danger"
                    )
                    return redirect(url_for("issues.index"))

                c.execute("""
                    INSERT INTO issue_register
                    (
                        issue_date,
                        employee_id,
                        item_id,
                        qty,
                        issued_by,
                        returnable,
                        return_due_date,
                        status,
                        remarks,
                        department
                    )
                    VALUES
                    (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    issue_date,
                    emp_id,
                    item_id,
                    qty,
                    issued_by,
                    returnable,
                    return_due,
                    "Issued",
                    remarks,
                    issue_department
                ))

                # Global stock
                c.execute("""
                    UPDATE items
                    SET stock = stock - %s
                    WHERE id=%s
                """, (qty, item_id))

                conn.commit()
                issued_count += 1

    except Exception as e:
        conn.rollback()
        flash(f"Error: {e}", "danger")

    finally:
        conn.close()

    if issued_count:
        flash(
            f"{issued_count} PPE/Equipment issue record(s) created successfully.",
            "success"
        )   

    return redirect(url_for("issues.index"))


@issues_bp.route('/issues/import-excel', methods=['POST'])
def import_excel():
    """
    Bulk-import issue records from an uploaded Excel file — same fields,
    same permission gate, and same department/stock validation as the
    single-record Add form above.

    Expected header row (any order, case-insensitive):
        Issue Date | Emp Code | Item Name | Qty | Returnable | Return Due Date | Remarks

    Issue Date and Return Due Date columns can be real Excel date cells
    or plain text in YYYY-MM-DD format. Returnable accepts Yes/No, 1/0,
    True/False, or can simply be left blank (treated as No).
    """
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    role = session.get('role')
    if not has_permission('can_create'):
        flash("You don't have permission to issue PPE/Equipment.", "danger")
        return redirect(url_for('issues.index'))

    file = request.files.get('excel_file')
    if not file or file.filename == '':
        flash('Please choose an Excel file to import.', 'danger')
        return redirect(url_for('issues.index'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Could not read the Excel file: {e}', 'danger')
        return redirect(url_for('issues.index'))

    header_row = [str(cell.value).strip().lower() if cell.value is not None else '' for cell in ws[1]]
    header_map = {h: idx for idx, h in enumerate(header_row) if h}

    required_cols = ['issue date', 'emp code','department','contractor','item name', 'qty']
    missing = [col for col in required_cols if col not in header_map]
    if missing:
        flash(
            "Excel is missing required column(s): " + ", ".join(missing) + ". "
            "Expected headers: Issue Date, Emp Code, Department, Contractor, Item Name, Qty, Returnable, Return Due Date, Remarks.",
            'danger'
        )
        return redirect(url_for('issues.index'))

    def _cell(row, col_name):
        idx = header_map.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx].value

    def _to_date_str(value):
        if value is None or value == '':
            return None
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        return str(value).strip()

    dept = session.get('department')
    is_admin = role in ['Admin', 'Super Admin']
    allowed_variants = None
    if not is_admin:
        display_name = _display_dept_name(dept)
        allowed_variants = [v.lower() for v in COMBINE_GROUPS.get(display_name, [dept])]

    conn = get_db()
    c = conn.cursor()
    imported = 0
    duplicates = 0
    skipped = 0
    errors = []

    try:
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            emp_code = _cell(row, 'emp code')
            item_name = _cell(row, 'item name')
            raw_qty = _cell(row, 'qty')
            raw_issue_date = _cell(row, 'issue date')
            raw_returnable = _cell(row, 'returnable')
            raw_return_due = _cell(row, 'return due date')
            raw_department = _cell(row, 'department')
            raw_contractor = _cell(row, 'contractor')
            remarks = _cell(row, 'remarks') or ''

            # Skip fully blank trailing rows silently (no error noise)
            if not any([emp_code, item_name, raw_qty, raw_issue_date]):
                continue

            if not emp_code or not item_name or raw_qty in (None, '') or not raw_issue_date:
                skipped += 1
                errors.append(f"Row {row_num}: missing Issue Date / Emp Code / Item Name / Qty — skipped.")
                continue

            emp_code = str(emp_code).strip()
            department = str(raw_department or '').strip()
            contractor = str(raw_contractor or '').strip()
            item_name = str(item_name).strip()

            try:
                qty = int(raw_qty)
                if qty <= 0:
                    raise ValueError()
            except (ValueError, TypeError):
                skipped += 1
                errors.append(f"Row {row_num}: invalid Qty '{raw_qty}' — skipped.")
                continue

            issue_date = _to_date_str(raw_issue_date)

            returnable = 1 if str(raw_returnable or '').strip().lower() in ('yes', '1', 'true') else 0
            return_due = _to_date_str(raw_return_due) if returnable else None

            c.execute("""
                    SELECT id, department, contractor
                    FROM employees
                    WHERE emp_code=%s
                    AND status='Active'
                """, (emp_code,))
            emp = fetchone(c)
            if not emp:
                skipped += 1
                errors.append(f"Row {row_num}: employee code '{emp_code}' not found (or inactive) — skipped.")
                continue
            # ---------- Department Validation ----------
            if department:
                excel_department = _display_dept_name(department)
                system_department = _display_dept_name(emp['department'] or '')

                if excel_department.lower() != system_department.lower():
                    skipped += 1
                    errors.append(
                        f"Row {row_num}: department mismatch "
                        f"(Excel: {department}, System: {emp['department']}) - skipped."
                    )
                    continue

            # ---------- Contractor Validation ----------
            if contractor:
                system_contractor = (emp['contractor'] or '').strip()

                if system_contractor.lower() != contractor.lower():
                    skipped += 1
                    errors.append(
                        f"Row {row_num}: contractor mismatch "
                        f"(Excel: {contractor}, System: {system_contractor}) - skipped."
                    )
                    continue

            if not is_admin and (emp['department'] or '').lower() not in allowed_variants:
                skipped += 1
                errors.append(f"Row {row_num}: employee '{emp_code}' is outside your department — skipped.")
                continue

            c.execute("SELECT id, stock FROM items WHERE LOWER(item_name)=LOWER(%s)", (item_name,))
            item = fetchone(c)
            if not item:
                skipped += 1
                errors.append(f"Row {row_num}: item '{item_name}' not found — skipped.")
                continue

            # ── DUPLICATE CHECK ──────────────────────────────────────────
            # Duplicate = Employee + Item + Qty + Issue Date + Returnable
            # all match an existing row exactly. Remarks and Return Due
            # Date are intentionally ignored for this check.
            c.execute("""
                SELECT id FROM issue_register
                WHERE employee_id=%s
                  AND item_id=%s
                  AND qty=%s
                  AND issue_date=%s
                  AND returnable=%s
                LIMIT 1
            """, (emp['id'], item['id'], qty, issue_date, returnable))
            existing = fetchone(c)
            if existing:
                duplicates += 1
                continue  # identical row already exists — silently skip

            if item['stock'] < qty:
                skipped += 1
                errors.append(f"Row {row_num}: insufficient overall stock for '{item_name}' — skipped.")
                continue

            employee_department = emp['department']
            issue_department = find_issue_department(c, item['id'], employee_department, qty)
            if not issue_department:
                skipped += 1
                errors.append(
                    f"Row {row_num}: insufficient stock in '{employee_department}' "
                    f"(or related departments) for '{item_name}' — skipped."
                )
                continue

            c.execute("""
                INSERT INTO issue_register
                (issue_date, employee_id, item_id, qty, issued_by, returnable,
                 return_due_date, status, remarks, department)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                issue_date, emp['id'], item['id'], qty, session['full_name'],
                returnable, return_due, 'Issued', remarks, issue_department
            ))
            c.execute("UPDATE items SET stock = stock - %s WHERE id=%s", (qty, item['id']))
            conn.commit()
            imported += 1

    except Exception as e:
        conn.rollback()
        flash(f'Import failed partway through: {e}', 'danger')
    finally:
        conn.close()

    flash(
        f'Import complete: {imported} new record(s) issued, '
        f'{duplicates} duplicate(s) skipped (already existed), '
        f'{skipped} row(s) skipped due to errors.',
        'success' if imported else 'warning'
    )
    if errors:
        shown = errors[:15]
        extra = f' … and {len(errors) - 15} more row(s) with issues.' if len(errors) > 15 else ''
        flash('Details: ' + ' | '.join(shown) + extra, 'warning')

    return redirect(url_for('issues.index'))


@issues_bp.route('/issues/edit/<int:id>', methods=['POST'])
def edit(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if not has_permission('can_edit'):                          # ← FIXED
        flash("You don't have permission to edit issue records.", "danger")
        return redirect(url_for('issues.index'))

    conn = get_db(); c = conn.cursor()
    try:
        c.execute("SELECT * FROM issue_register WHERE id=%s", (id,))
        old = fetchone(c)
        if not old:
            flash('Issue record not found.', 'danger')
            conn.close()
            return redirect(url_for('issues.index'))

        new_qty = int(request.form['qty'])
        qty_diff = new_qty - old['qty']

        if qty_diff > 0:
            c.execute("SELECT stock FROM items WHERE id=%s", (old['item_id'],))
            stock = fetchone(c)
            if not stock or stock['stock'] < qty_diff:
                conn.close()
                flash('Insufficient stock for this update!', 'danger')
                return redirect(url_for('issues.index'))

        returnable = 1 if request.form.get('returnable') else 0
        return_due = request.form.get('return_due_date') if returnable else None

        c.execute("""
            UPDATE issue_register
            SET issue_date=%s, qty=%s, returnable=%s, return_due_date=%s, remarks=%s
            WHERE id=%s
        """, (request.form['issue_date'], new_qty, returnable, return_due,
              request.form.get('remarks', ''), id))

        c.execute("UPDATE items SET stock=stock-%s WHERE id=%s", (qty_diff, old['item_id']))

        conn.commit()
        flash('Issue record updated successfully.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('issues.index'))


@issues_bp.route('/issues/delete/<int:id>', methods=['POST'])
def delete(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if not has_permission('can_delete'):                        # ← FIXED
        flash("You don't have permission to delete issue records.", "danger")
        return redirect(url_for('issues.index'))

    conn = get_db(); c = conn.cursor()
    try:
        c.execute("SELECT * FROM issue_register WHERE id=%s", (id,))
        row = fetchone(c)
        if not row:
            flash('Issue record not found.', 'danger')
            conn.close()
            return redirect(url_for('issues.index'))

        c.execute("DELETE FROM issue_register WHERE id=%s", (id,))
        c.execute("UPDATE items SET stock=stock+%s WHERE id=%s", (row['qty'], row['item_id']))
        conn.commit()
        flash('Issue record deleted and stock restored.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('issues.index'))


@issues_bp.route('/issues/download')
def download():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    # Download = read-only report; login gate hi purese ahe, extra permission nako vatly tar hi line kadhu shakta

    conn = get_db(); c = conn.cursor()
    dept = session.get('department')
    role = session.get('role')
    is_admin = role in ['Admin', 'Super Admin']

    if not is_admin and dept:
        display_name = _display_dept_name(dept)
        dept_variants = [v.lower() for v in COMBINE_GROUPS.get(display_name, [dept])]
    else:
        dept_variants = []

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    query = """
        SELECT ir.issue_date, e.emp_code, e.name as emp_name, e.department,
               i.item_name, ir.qty, i.unit, ir.status, ir.returnable,
               ir.return_due_date, ir.issued_by, ir.remarks
        FROM issue_register ir
        JOIN employees e ON ir.employee_id=e.id
        JOIN items i ON ir.item_id=i.id
        WHERE 1=1
    """
    params = []
    if dept_variants:
        query += " AND LOWER(e.department)=ANY(%s)"
        params.append(dept_variants)
    if from_date:
        query += " AND ir.issue_date >= %s"
        params.append(from_date)
    if to_date:
        query += " AND ir.issue_date <= %s"
        params.append(to_date)
    query += " ORDER BY ir.issue_date DESC"

    c.execute(query, tuple(params))
    rows = fetchall(c)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'PPE Issue Report'

    headers = ['Date', 'Emp Code', 'Name', 'Department', 'Item', 'Qty', 'Unit',
               'Status', 'Returnable', 'Return Due Date', 'Issued By', 'Remarks']

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, f'PPE / Equipment Issue Report ({from_date or "All"} to {to_date or "All"})')
    title_cell.font = Font(bold=True, size=13, color='2C3E50')
    title_cell.alignment = _ctr
    title_cell.fill = PatternFill('solid', fgColor='F6F8FB')
    for col in range(2, len(headers) + 1):
        ws.cell(1, col).fill = PatternFill('solid', fgColor='F6F8FB')

    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(2, idx, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = _ctr
        cell.border = _border

    row_idx = 3
    for r in rows:
        values = [
            r['issue_date'], r['emp_code'], r['emp_name'], r['department'],
            r['item_name'], r['qty'], r['unit'], r['status'],
            'Yes' if r['returnable'] else 'No',
            r['return_due_date'] or '', r['issued_by'], r['remarks'] or '',
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.border = _border
            cell.alignment = _left if col_idx in (3, 5, 12) else _ctr
        row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['L'].width = 26
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'PPE_Issue_Report_{from_date or "all"}_to_{to_date or "all"}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )