from flask import Blueprint, render_template, request, redirect, url_for, session, flash, jsonify, Response
from database.db import get_db, fetchall, fetchone
import io

items_bp = Blueprint('items', __name__)

VALID_CATEGORIES = ["PPE", "Operational Equipment", "Life Safety", "Gas Detection", "Fire Safety"]

@items_bp.route('/items')
def index():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM items ORDER BY category, item_name")
    items = fetchall(c)
    conn.close()
    return render_template('items.html', items=items)

@items_bp.route('/items/add', methods=['POST'])
def add():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    try:
        c.execute("INSERT INTO items (item_code,item_name,category,unit,min_stock,reorder_level,has_expiry,has_calibration) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                  (request.form['item_code'], request.form['item_name'], request.form['category'],
                   request.form.get('unit','Nos'), int(request.form.get('min_stock',0)),
                   int(request.form.get('reorder_level',0)),
                   1 if request.form.get('has_expiry') else 0,
                   1 if request.form.get('has_calibration') else 0))
        conn.commit()
        flash('Item added successfully.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('items.index'))

@items_bp.route('/items/edit/<int:id>', methods=['POST'])
def edit(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    c.execute("UPDATE items SET item_code=%s,item_name=%s,category=%s,unit=%s,min_stock=%s,reorder_level=%s,has_expiry=%s,has_calibration=%s WHERE id=%s",
              (request.form['item_code'], request.form['item_name'], request.form['category'],
               request.form.get('unit','Nos'), int(request.form.get('min_stock',0)),
               int(request.form.get('reorder_level',0)),
               1 if request.form.get('has_expiry') else 0,
               1 if request.form.get('has_calibration') else 0, id))
    conn.commit(); conn.close()
    flash('Item updated.', 'success')
    return redirect(url_for('items.index'))

@items_bp.route('/items/delete/<int:id>', methods=['POST'])
def delete(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    try:
        c.execute("DELETE FROM items WHERE id=%s", (id,))
        conn.commit()
        flash('Item deleted.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('items.index'))

@items_bp.route('/items/search')
def search():
    q = request.args.get('q','')
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT id, item_code, item_name, category, unit, stock FROM items WHERE item_name ILIKE %s OR item_code ILIKE %s LIMIT 20",
              (f'%{q}%', f'%{q}%'))
    results = fetchall(c)
    conn.close()
    return jsonify(results)


# ───────────────────────── EXCEL EXPORT (template / backup) ─────────────────────────
@items_bp.route('/items/export-excel')
def export_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('openpyxl is not installed.', 'danger')
        return redirect(url_for('items.index'))

    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM items ORDER BY category, item_name")
    items = fetchall(c)
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Items"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1A3A5C")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    headers = ["Item Code", "Item Name", "Category", "Unit", "Min Stock", "Reorder Level", "Has Expiry", "Has Calibration"]
    col_widths = [14, 30, 22, 10, 12, 14, 12, 15]
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center_align; cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    for row_idx, item in enumerate(items, start=2):
        row_data = [
            item.get("item_code"), item.get("item_name"), item.get("category", ""),
            item.get("unit", ""), item.get("min_stock", 0), item.get("reorder_level", 0),
            "Yes" if item.get("has_expiry") else "No",
            "Yes" if item.get("has_calibration") else "No",
        ]
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output); output.seek(0)
    return Response(output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items.xlsx"})


# ───────────────────────── EXCEL IMPORT (upsert + validation) ─────────────────────────
@items_bp.route('/items/import-excel', methods=['POST'])
def import_excel():
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if 'excel_file' not in request.files or request.files['excel_file'].filename == '':
        flash('No file selected.', 'danger')
        return redirect(url_for('items.index'))

    file = request.files['excel_file']
    filename = file.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash('Invalid file type. Please upload a .xlsx or .xls file.', 'danger')
        return redirect(url_for('items.index'))

    try:
        import openpyxl
    except ImportError:
        flash('openpyxl not installed.', 'danger')
        return redirect(url_for('items.index'))

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active

        # Header names matched case-insensitively / regardless of spacing
        HEADER_MAP = {
            'item code': 'item_code', 'itemcode': 'item_code', 'code': 'item_code', 'item_code': 'item_code',
            'item name': 'item_name', 'itemname': 'item_name', 'name': 'item_name', 'item_name': 'item_name',
            'category': 'category',
            'unit': 'unit', 'uom': 'unit',
            'min stock': 'min_stock', 'minstock': 'min_stock', 'min_stock': 'min_stock', 'minimum stock': 'min_stock',
            'reorder level': 'reorder_level', 'reorder': 'reorder_level', 'reorder_level': 'reorder_level',
            'has expiry': 'has_expiry', 'expiry': 'has_expiry', 'has_expiry': 'has_expiry',
            'has calibration': 'has_calibration', 'calibration': 'has_calibration', 'has_calibration': 'has_calibration',
        }

        first_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        col_map = {}
        for idx, cell_val in enumerate(first_row):
            if cell_val is None:
                continue
            key = str(cell_val).strip().lower()
            if key in HEADER_MAP and HEADER_MAP[key] not in col_map:
                col_map[HEADER_MAP[key]] = idx

        has_headers = 'item_code' in col_map and 'item_name' in col_map
        start_row = 2 if has_headers else 1

        POSITIONAL = {'item_code': 0, 'item_name': 1, 'category': 2, 'unit': 3,
                      'min_stock': 4, 'reorder_level': 5, 'has_expiry': 6, 'has_calibration': 7}

        def get_val(row_values, field):
            mapping = col_map if has_headers else POSITIONAL
            idx = mapping.get(field)
            if idx is None or idx >= len(row_values):
                return ''
            v = row_values[idx]
            return str(v).strip() if v is not None else ''

        # Case-insensitive lookup for category against the allowed list
        CATEGORY_LOOKUP = {cat.lower(): cat for cat in VALID_CATEGORIES}

        # Accepts Yes/No, True/False, 1/0, Y/N — case-insensitive — for checkbox-style fields
        TRUE_VALUES = {'yes', 'y', 'true', '1'}
        FALSE_VALUES = {'no', 'n', 'false', '0', ''}

        def parse_bool(raw, field_name, row_idx, errors):
            v = raw.strip().lower()
            if v in TRUE_VALUES:
                return 1
            if v in FALSE_VALUES:
                return 0
            errors.append(f"Row {row_idx}: invalid value '{raw}' for {field_name} (expected Yes/No)")
            return 0

        def parse_int(raw, field_name, row_idx, errors, default=0):
            if raw == '':
                return default
            try:
                return int(float(raw))
            except ValueError:
                errors.append(f"Row {row_idx}: invalid number '{raw}' for {field_name}")
                return default

        conn = get_db(); c = conn.cursor()
        added = updated = unchanged = skipped = 0
        error_log = []
        UPDATABLE_FIELDS = ['item_name', 'category', 'unit', 'min_stock', 'reorder_level', 'has_expiry', 'has_calibration']

        for row_idx, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            if not any(row):
                continue
            row_values = list(row)

            item_code = get_val(row_values, 'item_code')
            item_name = get_val(row_values, 'item_name')

            # item_code is the required key; name required only for brand-new items
            if not item_code:
                skipped += 1
                error_log.append(f"Row {row_idx}: missing Item Code — skipped")
                continue

            raw_category = get_val(row_values, 'category')
            category = CATEGORY_LOOKUP.get(raw_category.lower(), '') if raw_category else ''
            if raw_category and not category:
                error_log.append(f"Row {row_idx}: unknown category '{raw_category}' — left blank")

            unit = get_val(row_values, 'unit') or 'Nos'
            min_stock = parse_int(get_val(row_values, 'min_stock'), 'Min Stock', row_idx, error_log)
            reorder_level = parse_int(get_val(row_values, 'reorder_level'), 'Reorder Level', row_idx, error_log)
            has_expiry = parse_bool(get_val(row_values, 'has_expiry') or 'no', 'Has Expiry', row_idx, error_log)
            has_calibration = parse_bool(get_val(row_values, 'has_calibration') or 'no', 'Has Calibration', row_idx, error_log)

            incoming = {
                'item_name': item_name,
                'category': category,
                'unit': unit,
                'min_stock': min_stock,
                'reorder_level': reorder_level,
                'has_expiry': has_expiry,
                'has_calibration': has_calibration,
            }

            try:
                c.execute("SELECT * FROM items WHERE item_code=%s", (item_code,))
                existing = fetchone(c)

                if existing is None:
                    if not item_name:
                        skipped += 1
                        error_log.append(f"Row {row_idx}: new item '{item_code}' missing Item Name — skipped")
                        continue
                    c.execute("""
                        INSERT INTO items (item_code, item_name, category, unit, min_stock, reorder_level, has_expiry, has_calibration)
                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    """, (item_code, incoming['item_name'], incoming['category'], incoming['unit'],
                          incoming['min_stock'], incoming['reorder_level'],
                          incoming['has_expiry'], incoming['has_calibration']))
                    added += 1
                else:
                    changed_fields = {}
                    for field in UPDATABLE_FIELDS:
                        new_val = incoming[field]
                        old_val = existing.get(field)
                        # Skip blank/empty incoming text values so they don't wipe existing data
                        if field in ('item_name', 'category', 'unit') and new_val in ('', None):
                            continue
                        if new_val != old_val and new_val not in ('', None):
                            changed_fields[field] = new_val
                        elif field in ('min_stock', 'reorder_level', 'has_expiry', 'has_calibration') and new_val != (old_val or 0):
                            changed_fields[field] = new_val

                    if changed_fields:
                        set_clause = ", ".join(f"{f}=%s" for f in changed_fields)
                        params = list(changed_fields.values()) + [item_code]
                        c.execute(f"UPDATE items SET {set_clause} WHERE item_code=%s", params)
                        updated += 1
                    else:
                        unchanged += 1

            except Exception as row_err:
                skipped += 1
                error_log.append(f"Row {row_idx}: {row_err}")

        conn.commit(); conn.close()

        summary = f'Import complete: {added} added, {updated} updated, {unchanged} unchanged, {skipped} skipped.'
        flash(summary, 'success')
        if error_log:
            # show up to first 10 issues so the modal doesn't get flooded
            flash('Issues found: ' + ' | '.join(error_log[:10]) + (' ...' if len(error_log) > 10 else ''), 'warning')

    except Exception as e:
        flash(f'Failed to read Excel: {e}', 'danger')

    return redirect(url_for('items.index'))