from flask import Blueprint, render_template, request, redirect, url_for, session, flash, Response
from database.db import get_db, fetchall, fetchone
from datetime import date
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from modules.user_admin import has_permission   # ← ADDED

contractor_issues_bp = Blueprint('contractor_issues', __name__)
_thin = Side(style='thin', color='C7CDD4')
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left', vertical='center', wrap_text=True)


@contractor_issues_bp.route('/contractor-issues')
def index():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    dept = session.get('department')

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    role = session.get("role")
    is_admin = role in ["Admin", "Super Admin"]

    # Auto-sync contractors from active employees into contractors table if missing
    c.execute("""
        INSERT INTO contractors (name)
        SELECT DISTINCT TRIM(e.contractor)
        FROM employees e
        WHERE e.contractor IS NOT NULL AND TRIM(e.contractor) <> ''
          AND NOT EXISTS (
              SELECT 1 FROM contractors c WHERE UPPER(TRIM(c.name)) = UPPER(TRIM(e.contractor))
          )
    """)
    conn.commit()

    if is_admin:
        c.execute("""
            SELECT DISTINCT
                c.id,
                c.name,
                COALESCE(c.contact, '') AS contact,
                e.department
            FROM employees e
            JOIN contractors c
                ON UPPER(TRIM(c.name)) = UPPER(TRIM(e.contractor))
            WHERE
                e.status = 'Active'
                AND e.contractor IS NOT NULL
                AND TRIM(e.contractor) <> ''
            ORDER BY e.department, c.name
        """)
    else:
        c.execute("""
            SELECT DISTINCT
                c.id,
                c.name,
                COALESCE(c.contact, '') AS contact,
                e.department
            FROM employees e
            JOIN contractors c
                ON UPPER(TRIM(c.name)) = UPPER(TRIM(e.contractor))
            WHERE
                e.status = 'Active'
                AND LOWER(TRIM(e.department)) = LOWER(TRIM(%s))
                AND e.contractor IS NOT NULL
                AND TRIM(e.contractor) <> ''
            ORDER BY c.name
        """, (dept,))

    contractors_raw = fetchall(c)

    contractors = [
        {
            "id": ct["id"],
            "name": ct["name"],
            "label": f"{ct['name']} ({ct['department']})" if is_admin else ct["name"]
        }
        for ct in contractors_raw
    ]


    query = """
    SELECT
        e.id,
        e.emp_code,
        e.name,
        c.id AS contractor_id
    FROM employees e
    JOIN contractors c
        ON UPPER(TRIM(c.name)) = UPPER(TRIM(e.contractor))
    WHERE
        e.status='Active'
        AND e.contractor IS NOT NULL
        AND TRIM(e.contractor) <> ''
    """

    params = []

    if not is_admin and dept:
        query += " AND LOWER(TRIM(e.department)) = LOWER(TRIM(%s))"
        params.append(dept)

    query += " ORDER BY e.name"

    c.execute(query, tuple(params))
    contractor_employees_raw = fetchall(c)

    contractor_employees = []

    for emp in contractor_employees_raw:
        contractor_employees.append({
            "id": emp["id"],
            "contractor_id": emp["contractor_id"],
            "label": f"{emp['emp_code']} - {emp['name']}"
        })

    c.execute("SELECT id, item_name, unit FROM items ORDER BY item_name")
    items_raw = fetchall(c)
    items = [{
        "id": i["id"],
        "unit": i["unit"],
        "label": i["item_name"],
    } for i in items_raw]

    query = """
        SELECT
            cir.id,
            cir.contractor_id,
            cir.employee_id,
            cir.item_id,
            cir.issue_date,
            cir.qty,
            cir.returnable,
            cir.return_due_date,
            cir.status,
            cir.issued_by,
            cir.remarks,

            ct.name AS contractor_name,
            e.department,

            e.id AS emp_id,
            e.emp_code,
            e.name AS employee_name,

            i.item_name,
            i.unit

        FROM contractor_issue_register cir

        LEFT JOIN contractors ct
            ON ct.id = cir.contractor_id

        LEFT JOIN employees e
            ON e.id = cir.employee_id

        LEFT JOIN items i
            ON i.id = cir.item_id

        WHERE 1=1
        """
    params = []
    if not is_admin and dept:
        query += " AND LOWER(TRIM(e.department)) = LOWER(TRIM(%s))"
        params.append(dept)
    if from_date:
        query += " AND cir.issue_date >= %s"; params.append(from_date)
    if to_date:
        query += " AND cir.issue_date <= %s"; params.append(to_date)
    query += " ORDER BY cir.issue_date DESC"

    c.execute(query, tuple(params))
    issues = fetchall(c)
    conn.close()

    can_create = has_permission('can_create')
    can_edit   = has_permission('can_edit')
    can_delete = has_permission('can_delete')

    return render_template('contractor_issues.html', contractors=contractors, items=items, issues=issues,
                            contractor_employees=contractor_employees,
                            today=date.today(), from_date=from_date, to_date=to_date,
                            can_create=can_create, can_edit=can_edit, can_delete=can_delete)


@contractor_issues_bp.route('/contractor-issues/add', methods=['POST'])
def add():
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if not has_permission('can_create'):
        flash("You don't have permission to issue PPE to contractors.", "danger")
        return redirect(url_for("contractor_issues.index"))

    conn = get_db()
    c = conn.cursor()

    try:
        contractor_id = int(request.form['contractor_id'])
        employee_ids = [int(x) for x in request.form['employee_id'].split(',') if x.strip()]
        item_ids = [int(x) for x in request.form['item_id'].split(',') if x.strip()]

        qty = int(request.form['qty'])
        issue_date = request.form['issue_date']
        remarks = request.form.get('remarks', '')
        issued_by = session['full_name']

        returnable = 1 if request.form.get('returnable') else 0
        return_due = request.form.get('return_due_date') if returnable else None

        role = session.get("role")
        dept = session.get("department")
        is_admin = role in ["Admin", "Super Admin"]

        # Department validation: for non-admins, ensure all selected employees belong to the logged-in department
        if not is_admin and dept:
            session_dept = dept.strip().lower()
            for emp_id in employee_ids:
                c.execute("SELECT department FROM employees WHERE id=%s", (emp_id,))
                emp_row = fetchone(c)
                if not emp_row or (emp_row["department"] or "").strip().lower() != session_dept:
                    conn.close()
                    flash("You can only issue PPE to employees in your department.", "danger")
                    return redirect(url_for("contractor_issues.index"))

        for emp_id in employee_ids:
            for item_id in item_ids:
                c.execute("""
                    INSERT INTO contractor_issue_register
                    (issue_date, contractor_id, employee_id, item_id, qty, issued_by,
                     returnable, return_due_date, status, remarks)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (issue_date, contractor_id, emp_id, item_id, qty, issued_by,
                      returnable, return_due, 'Issued', remarks))

        conn.commit()
        flash("PPE/Equipment issued to contractor employee(s) successfully.", "success")

    except Exception as e:
        import traceback
        traceback.print_exc()
        conn.rollback()
        flash(f"Error: {e}", "danger")
    finally:
        conn.close()

    return redirect(url_for("contractor_issues.index"))


@contractor_issues_bp.route('/contractor-issues/edit/<int:id>', methods=['POST'])
def edit(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if not has_permission('can_edit'):                          # ← FIXED
        flash("You don't have permission to edit contractor issue records.", "danger")
        return redirect(url_for('contractor_issues.index'))

    conn = get_db(); c = conn.cursor()
    try:
        c.execute("SELECT * FROM contractor_issue_register WHERE id=%s", (id,))
        old = fetchone(c)
        if not old:
            flash('Issue record not found.', 'danger')
            conn.close()
            return redirect(url_for('contractor_issues.index'))

        contractor_id = request.form.get('contractor_id')
        employee_id = request.form.get('employee_id')
        item_id = request.form.get('item_id')
        issue_date = request.form.get('issue_date')
        qty_raw = request.form.get('qty')

        def is_blank(v):
            return not v or v.strip().lower() in ('none', 'null')

        if is_blank(employee_id) and old.get('employee_id'):
            employee_id = old['employee_id']

        missing = [name for name, val in [
            ('Contractor', contractor_id), ('Employee', employee_id),
            ('Item', item_id), ('Issue Date', issue_date), ('Quantity', qty_raw)
        ] if is_blank(val)]
        if missing:
            conn.close()
            flash(f"Update failed — missing required field(s): {', '.join(missing)}. "
                  f"Please reselect them and try again.", 'danger')
            return redirect(url_for('contractor_issues.index'))

        try:
            contractor_id = int(contractor_id)
            employee_id = int(employee_id)
            item_id = int(item_id)
            new_qty = int(qty_raw)
        except (TypeError, ValueError):
            conn.close()
            flash("Update failed — Contractor, Employee, Item, and Quantity must be valid numbers.", 'danger')
            return redirect(url_for('contractor_issues.index'))

        returnable = 1 if request.form.get('returnable') else 0
        return_due = request.form.get('return_due_date') if returnable else None

        c.execute("""
            UPDATE contractor_issue_register
            SET contractor_id=%s,
                employee_id=%s,
                item_id=%s,
                issue_date=%s,
                qty=%s,
                returnable=%s,
                return_due_date=%s,
                remarks=%s
            WHERE id=%s
        """, (
            contractor_id, employee_id, item_id, issue_date,
            new_qty, returnable, return_due,
            request.form.get("remarks", ""), id
        ))

        if c.rowcount == 0:
            conn.rollback()
            flash("Update failed — no matching record found to update.", 'danger')
            conn.close()
            return redirect(url_for('contractor_issues.index'))

        conn.commit()
        flash('Issue record updated successfully.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('contractor_issues.index'))


@contractor_issues_bp.route('/contractor-issues/delete/<int:id>', methods=['POST'])
def delete(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    if not has_permission('can_delete'):                        # ← FIXED
        flash("You don't have permission to delete contractor issue records.", "danger")
        return redirect(url_for('contractor_issues.index'))

    conn = get_db(); c = conn.cursor()
    try:
        c.execute("SELECT * FROM contractor_issue_register WHERE id=%s", (id,))
        row = fetchone(c)
        if not row:
            flash('Issue record not found.', 'danger')
            conn.close()
            return redirect(url_for('contractor_issues.index'))

        c.execute("DELETE FROM contractor_issue_register WHERE id=%s", (id,))
        conn.commit()
        flash('Issue record deleted.', 'success')
    except Exception as e:
        conn.rollback(); flash(f'Error: {e}', 'danger')
    conn.close()
    return redirect(url_for('contractor_issues.index'))


@contractor_issues_bp.route('/contractor-issues/download')
def download():
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    conn = get_db(); c = conn.cursor()
    dept = session.get('department')
    role = session.get("role")
    is_admin = role in ["Admin", "Super Admin"]

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    query = """
        SELECT cir.issue_date, ct.name as contractor_name, ct.department,
               i.item_name, cir.qty, i.unit, cir.status, cir.returnable,
               cir.return_due_date, cir.issued_by, cir.remarks
        FROM contractor_issue_register cir
        JOIN contractors ct ON cir.contractor_id=ct.id
        JOIN items i ON cir.item_id=i.id
        WHERE 1=1
    """
    params = []
    if not is_admin:
        query += " AND ct.department=%s"
        params.append(dept)
    if from_date:
        query += " AND cir.issue_date >= %s"
        params.append(from_date)
    if to_date:
        query += " AND cir.issue_date <= %s"
        params.append(to_date)
    query += " ORDER BY cir.issue_date DESC"

    c.execute(query, tuple(params))
    rows = fetchall(c)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contractor PPE Issue Report'

    headers = ['Date', 'Contractor', 'Department', 'Item', 'Qty', 'Unit',
               'Status', 'Returnable', 'Return Due Date', 'Issued By', 'Remarks']

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, f'Contractor PPE / Equipment Issue Report ({from_date or "All"} to {to_date or "All"})')
    title_cell.font = Font(bold=True, size=13, color='2C3E50')
    title_cell.alignment = _ctr
    title_cell.fill = PatternFill('solid', fgColor='F6F8FB')
    for col in range(2, len(headers) + 1):
        ws.cell(1, col).fill = PatternFill('solid', fgColor='F6F8FB')

    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(2, idx, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = _ctr
        cell.border = _border

    row_idx = 3
    for r in rows:
        values = [
            r['issue_date'], r['contractor_name'], r['department'] or '',
            r['item_name'], r['qty'], r['unit'], r['status'],
            'Yes' if r['returnable'] else 'No',
            r['return_due_date'] or '', r['issued_by'], r['remarks'] or '',
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.border = _border
            cell.alignment = _left if col_idx in (2, 4, 11) else _ctr
        row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['K'].width = 26
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'Contractor_PPE_Issue_Report_{from_date or "all"}_to_{to_date or "all"}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )

# ─────────────────────── IMPORT EXCEL ───────────────────────
@contractor_issues_bp.route('/contractor-issues/import-excel', methods=['POST'])
def import_excel():
    """
    Bulk-import contractor issue records from an uploaded Excel file.

    Expected header row (any order, case-insensitive):
        Issue Date | Emp Code | Name | Department | Contractor | Item Name |
        Qty | Returnable | Return Due Date | Remarks

    'Name' column is accepted but treated as informational only — the
    employee is always resolved by Emp Code.
    Returnable: Yes/No/1/0/True/False (blank = No).
    """
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_create'):
        flash("You don't have permission to issue PPE/Equipment.", 'danger')
        return redirect(url_for('contractor_issues.index'))

    file = request.files.get('excel_file')
    if not file or file.filename == '':
        flash('Please choose an Excel file to import.', 'danger')
        return redirect(url_for('contractor_issues.index'))

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Could not read the Excel file: {e}', 'danger')
        return redirect(url_for('contractor_issues.index'))

    header_row = [
        str(cell.value).strip().lower() if cell.value is not None else ''
        for cell in ws[1]
    ]
    header_map = {h: idx for idx, h in enumerate(header_row) if h}

    required_cols = ['issue date', 'emp code', 'department', 'contractor', 'item name', 'qty']
    missing = [col for col in required_cols if col not in header_map]
    if missing:
        flash(
            "Excel is missing required column(s): " + ", ".join(missing) + ". "
            "Expected headers: Issue Date, Emp Code, Name, Department, Contractor, "
            "Item Name, Qty, Returnable, Return Due Date, Remarks.",
            'danger'
        )
        return redirect(url_for('contractor_issues.index'))

    def _cell(row, col_name):
        idx = header_map.get(col_name)
        if idx is None or idx >= len(row):
            return None
        return row[idx].value

    def _to_date_str(value):
        if value is None or value == '':
            return None
        if hasattr(value, 'strftime'):
            return value.strftime('%Y-%m-%d')
        return str(value).strip()

    role = session.get('role')
    dept = session.get('department')
    is_admin = role in ['Admin', 'Super Admin']

    conn = get_db()
    c = conn.cursor()
    imported = 0
    duplicates = 0
    skipped = 0
    errors = []

    try:
        for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
            emp_code       = _cell(row, 'emp code')
            item_name      = _cell(row, 'item name')
            raw_qty        = _cell(row, 'qty')
            raw_issue_date = _cell(row, 'issue date')
            raw_returnable = _cell(row, 'returnable')
            raw_return_due = _cell(row, 'return due date')
            raw_department = _cell(row, 'department')
            raw_contractor = _cell(row, 'contractor')
            remarks        = _cell(row, 'remarks') or ''

            # Skip fully blank trailing rows silently
            if not any([emp_code, item_name, raw_qty, raw_issue_date]):
                continue

            if not emp_code or not item_name or raw_qty in (None, '') or not raw_issue_date:
                skipped += 1
                errors.append(
                    f"Row {row_num}: missing Issue Date / Emp Code / Item Name / Qty -- skipped."
                )
                continue

            emp_code   = str(emp_code).strip()
            department = str(raw_department or '').strip()
            contractor = str(raw_contractor or '').strip()
            item_name  = str(item_name).strip()

            try:
                qty = int(raw_qty)
                if qty <= 0:
                    raise ValueError()
            except (ValueError, TypeError):
                skipped += 1
                errors.append(f"Row {row_num}: invalid Qty '{raw_qty}' -- skipped.")
                continue

            issue_date = _to_date_str(raw_issue_date)
            returnable = 1 if str(raw_returnable or '').strip().lower() in ('yes', '1', 'true') else 0
            return_due = _to_date_str(raw_return_due) if returnable else None

            # -- Look up employee ----------------------------------------
            c.execute("""
                SELECT id, department, contractor
                FROM employees
                WHERE emp_code=%s AND status='Active'
            """, (emp_code,))
            emp = fetchone(c)
            if not emp:
                skipped += 1
                errors.append(
                    f"Row {row_num}: employee code '{emp_code}' not found "
                    f"(or inactive) -- skipped."
                )
                continue

            # -- Department access check for non-admins -------------------
            if not is_admin:
                session_dept = (dept or '').strip().lower()
                emp_dept = (emp['department'] or '').strip().lower()
                if emp_dept != session_dept:
                    skipped += 1
                    errors.append(
                        f"Row {row_num}: employee '{emp_code}' is outside your "
                        f"department -- skipped."
                    )
                    continue

            # -- Contractor name validation --------------------------------
            if contractor:
                system_contractor = (emp['contractor'] or '').strip()
                if system_contractor.lower() != contractor.lower():
                    skipped += 1
                    errors.append(
                        f"Row {row_num}: contractor mismatch "
                        f"(Excel: {contractor}, System: {system_contractor}) -- skipped."
                    )
                    continue

            # -- Resolve contractor_id ------------------------------------
            ctr_name = (emp['contractor'] or contractor).strip()
            if not ctr_name:
                skipped += 1
                errors.append(
                    f"Row {row_num}: employee '{emp_code}' has no contractor "
                    f"assigned -- skipped."
                )
                continue

            c.execute(
                "SELECT id FROM contractors WHERE LOWER(TRIM(name))=LOWER(TRIM(%s))",
                (ctr_name,)
            )
            ctr_row = fetchone(c)
            if not ctr_row:
                c.execute(
                    "INSERT INTO contractors (name) VALUES (%s) RETURNING id",
                    (ctr_name,)
                )
                ctr_row = fetchone(c)
            contractor_id = ctr_row['id']

            # -- Resolve item ----------------------------------------------
            c.execute("SELECT id FROM items WHERE LOWER(TRIM(item_name))=LOWER(TRIM(%s))", (item_name,))
            item = fetchone(c)
            if not item:
                skipped += 1
                errors.append(f"Row {row_num}: item '{item_name}' not found -- skipped.")
                continue

            # -- Duplicate check: skip if all fields match an existing record ------------
            c.execute("""
                SELECT id FROM contractor_issue_register
                WHERE employee_id=%s
                  AND contractor_id=%s
                  AND item_id=%s
                  AND qty=%s
                  AND issue_date=%s
                  AND returnable=%s
                  AND COALESCE(return_due_date, '') = COALESCE(%s, '')
                  AND COALESCE(remarks, '') = COALESCE(%s, '')
                LIMIT 1
            """, (
                emp['id'], contractor_id, item['id'], qty, issue_date, returnable,
                return_due or '', remarks or ''
            ))
            if fetchone(c):
                duplicates += 1
                continue

            # -- Insert ---------------------------------------------------
            c.execute("""
                INSERT INTO contractor_issue_register
                (issue_date, contractor_id, employee_id, item_id, qty,
                 issued_by, returnable, return_due_date, status, remarks)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                issue_date, contractor_id, emp['id'], item['id'], qty,
                session['full_name'], returnable, return_due, 'Issued', remarks
            ))
            conn.commit()
            imported += 1

    except Exception as e:
        conn.rollback()
        flash(f'Import failed partway through: {e}', 'danger')
    finally:
        conn.close()

    flash(
        f'Import complete: {imported} new record(s) issued, '
        f'{duplicates} duplicate(s) skipped (already existed), '
        f'{skipped} row(s) skipped due to errors.',
        'success' if imported else 'warning'
    )
    if errors:
        shown = errors[:15]
        extra = f' and {len(errors) - 15} more row(s) with issues.' if len(errors) > 15 else ''
        flash('Details: ' + ' | '.join(shown) + extra, 'warning')

    return redirect(url_for('contractor_issues.index'))


# ─────────────────────── IMPORT TEMPLATE DOWNLOAD ───────────────────────
@contractor_issues_bp.route('/contractor-issues/import-template')
def download_import_template():
    """Returns a blank Excel template with the correct column headers and
    one example row so users know the expected format."""
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    wb = Workbook()
    ws = wb.active
    ws.title = 'Contractor Issue Import'

    headers = [
        'Issue Date', 'Emp Code', 'Name', 'Department', 'Contractor',
        'Item Name', 'Qty', 'Returnable', 'Return Due Date', 'Remarks'
    ]
    col_widths = [14, 14, 28, 20, 28, 24, 8, 12, 16, 40]

    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='1A3A5C')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(1, col_idx, h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = _border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    # One example row so users know what format is expected
    example = [
        '2026-08-01', 'CIPD003680', 'AAKASH CHANDRAKANT PATIL',
        'Operations', 'R J Enterprises', 'Helmet',
        2, 'Yes', '2026-09-01', 'Site work'
    ]
    eg_align = Alignment(horizontal='left', vertical='center')
    eg_fill  = PatternFill('solid', fgColor='EEF2FF')
    for col_idx, val in enumerate(example, start=1):
        cell = ws.cell(2, col_idx, val)
        cell.alignment = eg_align
        cell.fill      = eg_fill
        cell.border    = _border

    ws.freeze_panes = 'A2'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="contractor_issue_import_template.xlsx"'},
    )
