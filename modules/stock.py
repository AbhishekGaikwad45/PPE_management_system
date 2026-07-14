from flask import Blueprint, render_template, request, redirect, url_for, session, flash ,Response
from database.db import get_db, fetchall, fetchone
from datetime import date
from modules.user_admin import has_permission 
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
stock_bp = Blueprint('stock', __name__)
_thin = Side(style='thin', color='C7CDD4')
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_ctr = Alignment(horizontal='center', vertical='center', wrap_text=True)
_left = Alignment(horizontal='left', vertical='center', wrap_text=True)




# ── index() route — permission flags pass kele ─────────────────
@stock_bp.route('/stock')
def index():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM items ORDER BY item_name")
    items = fetchall(c)

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    query = """
        SELECT r.*, i.item_name, i.item_code
        FROM stock_receipts r
        JOIN items i ON r.item_id=i.id
        WHERE 1=1
    """
    params = []
    if from_date:
        query += " AND r.receipt_date >= %s"
        params.append(from_date)
    if to_date:
        query += " AND r.receipt_date <= %s"
        params.append(to_date)
    query += " ORDER BY r.receipt_date DESC"

    c.execute(query, tuple(params))
    receipts = fetchall(c)
    conn.close()

    can_create = has_permission('can_create')
    can_edit = has_permission('can_edit')
    can_delete = has_permission('can_delete')

    return render_template('stock.html', items=items, receipts=receipts, today=date.today(),
                            can_create=can_create, can_edit=can_edit, can_delete=can_delete,
                            from_date=from_date, to_date=to_date)


# ── receive() route — can_create check add kela ─────────────────
@stock_bp.route('/stock/receive', methods=['POST'])
def receive():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_create'):                       # ← ADD
        flash('You do not have permission to receive stock.', 'danger')
        return redirect(url_for('stock.index'))
    conn = get_db(); c = conn.cursor()

    item_id = request.form.get('item_id')
    receipt_date = request.form.get('receipt_date')
    grn_no = request.form.get('grn_no')
    qty = int(request.form.get('qty', 0))
    remarks = request.form.get('remarks')
    received_by = session.get('user')

    c.execute("""INSERT INTO stock_receipts (receipt_date, item_id, qty, grn_no, received_by, remarks)
                 VALUES (%s,%s,%s,%s,%s,%s)""",
              (receipt_date, item_id, qty, grn_no, received_by, remarks))
    c.execute("UPDATE items SET stock = stock + %s WHERE id=%s", (qty, item_id))

    conn.commit()
    conn.close()
    flash('Stock received successfully.', 'success')
    return redirect(url_for('stock.index'))


# ── edit() route — can_edit check add kela ──────────────────────
@stock_bp.route('/stock/edit/<int:id>', methods=['POST'])
def edit(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_edit'):                         # ← ADD
        flash('You do not have permission to edit receipts.', 'danger')
        return redirect(url_for('stock.index'))
    conn = get_db(); c = conn.cursor()

    c.execute("SELECT * FROM stock_receipts WHERE id=%s", (id,))
    old = fetchone(c)
    if not old:
        conn.close()
        flash('Receipt not found.', 'danger')
        return redirect(url_for('stock.index'))

    receipt_date = request.form.get('receipt_date')
    grn_no = request.form.get('grn_no')
    qty = int(request.form.get('qty', 0))
    remarks = request.form.get('remarks')
    qty_diff = qty - old['qty']

    c.execute("""UPDATE stock_receipts SET receipt_date=%s, grn_no=%s, qty=%s, remarks=%s
                 WHERE id=%s""", (receipt_date, grn_no, qty, remarks, id))
    c.execute("UPDATE items SET stock = stock + %s WHERE id=%s", (qty_diff, old['item_id']))

    conn.commit()
    conn.close()
    flash('Receipt updated successfully.', 'success')
    return redirect(url_for('stock.index'))


# ── delete() route — can_delete check add kela ──────────────────
@stock_bp.route('/stock/delete/<int:id>', methods=['POST'])
def delete(id):
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    if not has_permission('can_delete'):                       # ← ADD
        flash('You do not have permission to delete receipts.', 'danger')
        return redirect(url_for('stock.index'))
    conn = get_db(); c = conn.cursor()

    c.execute("SELECT * FROM stock_receipts WHERE id=%s", (id,))
    rec = fetchone(c)
    if not rec:
        conn.close()
        flash('Receipt not found.', 'danger')
        return redirect(url_for('stock.index'))

    c.execute("DELETE FROM stock_receipts WHERE id=%s", (id,))
    c.execute("UPDATE items SET stock = stock - %s WHERE id=%s", (rec['qty'], rec['item_id']))

    conn.commit()
    conn.close()
    flash('Receipt deleted successfully.', 'success')
    return redirect(url_for('stock.index'))

@stock_bp.route('/stock/ledger')
def ledger():
    if 'user' not in session:
        return redirect(url_for('auth.login'))
    conn = get_db(); c = conn.cursor()
    item_id = request.args.get('item_id')
    c.execute("SELECT * FROM items ORDER BY item_name")
    items = fetchall(c)
    ledger_data = []; selected_item = None
    if item_id:
        c.execute("SELECT * FROM items WHERE id=%s", (item_id,))
        selected_item = fetchone(c)
        c.execute("SELECT receipt_date as txn_date, qty, 'Receipt' as type, grn_no as ref FROM stock_receipts WHERE item_id=%s", (item_id,))
        receipts = fetchall(c)
        c.execute("SELECT i.issue_date as txn_date, i.qty, 'Issue' as type, e.name as ref FROM issue_register i JOIN employees e ON i.employee_id=e.id WHERE i.item_id=%s", (item_id,))
        issues = fetchall(c)
        ledger_data = sorted(receipts + issues, key=lambda x: x['txn_date'])
        running_balance = 0
        for row in ledger_data:
            if row['type'] == 'Receipt':
                running_balance += row['qty']; row['balance'] = running_balance
                row['in_qty'] = row['qty']; row['out_qty'] = ''
            else:
                running_balance -= row['qty']; row['balance'] = running_balance
                row['in_qty'] = ''; row['out_qty'] = row['qty']
    conn.close()
    return render_template('ledger.html', items=items, ledger_data=ledger_data, selected_item=selected_item, item_id=item_id)

@stock_bp.route('/stock/download')
def download():
    if 'user' not in session:
        return redirect(url_for('auth.login'))

    conn = get_db(); c = conn.cursor()

    from_date = request.args.get('from_date', '')
    to_date = request.args.get('to_date', '')

    query = """
        SELECT r.receipt_date, r.grn_no, i.item_name, i.item_code,
               r.qty, i.unit, r.received_by, r.remarks
        FROM stock_receipts r
        JOIN items i ON r.item_id=i.id
        WHERE 1=1
    """
    params = []
    if from_date:
        query += " AND r.receipt_date >= %s"
        params.append(from_date)
    if to_date:
        query += " AND r.receipt_date <= %s"
        params.append(to_date)
    query += " ORDER BY r.receipt_date DESC"

    c.execute(query, tuple(params))
    rows = fetchall(c)
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Receipts'

    headers = ['Date', 'GRN No', 'Item Name', 'Item Code', 'Qty', 'Unit', 'Received By', 'Remarks']

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(1, 1, f'Stock Receipt Report ({from_date or "All"} to {to_date or "All"})')
    title_cell.font = Font(bold=True, size=13, color='2C3E50')
    title_cell.alignment = _ctr
    title_cell.fill = PatternFill('solid', fgColor='F6F8FB')
    for col in range(2, len(headers) + 1):
        ws.cell(1, col).fill = PatternFill('solid', fgColor='F6F8FB')

    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(2, idx, h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='2C3E50')
        cell.alignment = _ctr
        cell.border = _border

    row_idx = 3
    for r in rows:
        values = [
            r['receipt_date'], r['grn_no'] or '', r['item_name'], r['item_code'] or '',
            r['qty'], r.get('unit', '') or '', r['received_by'] or '', r['remarks'] or '',
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row_idx, col_idx, val)
            cell.border = _border
            cell.alignment = _left if col_idx in (3, 8) else _ctr
        row_idx += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['H'].width = 28
    ws.freeze_panes = 'A3'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f'Stock_Receipt_Report_{from_date or "all"}_to_{to_date or "all"}.xlsx'
    return Response(
        buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{fname}"'},
    )